#! python3
# updateProduce.py - corrects costs in produce sales spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# Produce types and their updated prices
price_updates = {'Garlic':3.07,
                'Celery':1.19,
                'Lemon':1.27}

# Loop through the rows and update the prices
for rowNum in range(2,sheet.max_row+1):
    produceName = sheet.cell(row=rowNum,column = 1).value
    if produceName in price_updates:
        sheet.cell(row = rowNum,column = 2).value = price_updates[produceName]

wb.save('updatedProduceSales.xlsx')