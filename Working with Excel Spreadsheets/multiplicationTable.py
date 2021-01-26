#! python3
# multiplicationTable.py - takes user input and creates multiplied values

import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
n = int(input('Enter Number N: '))
for i in range(1,n+1):
    sheet.cell(1,i+1).value = i
    sheet.cell(i+1,1).value = i
for i in range(2,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
        sheet.cell(i,j).value = sheet.cell(1,j).value*sheet.cell(i,1).value
wb.save('multiplicationTable.xlsx')
